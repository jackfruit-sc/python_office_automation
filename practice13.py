from openpyxl import load_workbook, Workbook
import glob

#批量查询 excel 数据
goal_dir = '/home/shiyanlou/exp4'
new_workbook = Workbook()
new_sheet = new_workbook.active

flag = 0
count = 1

for file in glob.glob(goal_dir + '/*.xlsx'):
	print('处理第 {} 个xlsx'.str(count))
	count += 1

	workbook = load_workbook(file)
	sheet = workbook.active

	buy_mount = sheet['F']

	row_lst = []

	for cell in buy_mount:
		if isinstance(cell.value, int) and cell.value > 50:
			# print(cell.row)
			row_lst.append(cell.row)


	if not flag:
		header = sheet[1]
		header_lst = []

		for cell in header:
			header_lst.append(cell.value)

		new_sheet.append(header_lst)
		flag = 1

	for row in row_lst:
		data_lst = []
		for cell in sheet[row]:
			data_lst.append(cell.value)
		new_sheet.append(data_lst)

	new_workbook.save(goal_dir + '/' + 'new_baby_trade.xlsx')